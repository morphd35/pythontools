import pandas as pd 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import re

def generate_business_name(api_element, existing_name):
    if pd.notna(existing_name) and existing_name.strip().upper != 'N/A':
        return existing_name.strip()
    
    # Convert camel case or PascalCase to space-seperated words
    words = re.findall(f'[A-Z]?[a-z]+|[A-Z]+(?![a-z])', api_element)
    return ' '.join(word.capitalize() for word in words)

def generate_possible_values_list(xlsx_path: str):
    # Load the workbook and the RD tab
    wb = load_workbook(xlsx_path)
    if "Resource Details" not in wb.sheetnames:
        raise ValueError("Resource Details tab not found in workbook")
    rd_df = pd.read_excel(xlsx_path, sheet_name="Resource Details", header=1) # header row 2

    # Check required columns
    required_cols = ["Business Name", "Possible Values"]
    for col in required_cols:
        if col not in rd_df.columns:
            raise ValueError(f"Required column '{col}' missing in Resource Details")
        
    # Prepare list of rows for PVL
    pvl_rows = []

    for _, row in rd_df.iterrows():
        business_name = row["Business Name"]
        if pd.isna(business_name) or str(business_name).strip().upper() == "N/A":
            # Optimally convert API element to business-like name
            api_element = row.get("API Element", "")
            if api_element:
                business_name = convert_api_element_to_name(api_element)
            else:
                business_name = "UNKNOWN"

        possible_values_cell = row["Possible Values"]
        if pd.isna(possible_values_cell):
            continue

        # Split on new lines
        values = str(possible_values_cell).splitlines()
        for val in values:
            val = val.strip()
            if not val:
                continue


            # Check if there is a description using ' - '
            if ' - ' in val:
                value_part, desc_part = map(str.strip, val.split(' - ', 1))
            else:
                value_part = val
                desc_part = f"Identifies {business_name.strip().title()} as {value_part.strip().title()}." # default description

            pvl_rows.append({
                "Business Name": business_name,
                "Possible Value": value_part,
                "Possible Value Description": desc_part


            })    

    # Convert to dataframe
    pvl_df = pd.DataFrame(pvl_rows)

    # Normalize for dedupe: lowercase for compare
    pvl_df["dup_key"] = (pvl_df["Business Name"].str.lower() + '|' +
                         pvl_df["Possible Value"].str.lower() + '|' +
                         pvl_df["Possible Value Description"].str.lower()
                        )

    # Remove duplicates based on Business Name, Possible Value and Possible Value description.
    pvl_df = pvl_df.drop_duplicates(subset=['dup_key']).drop(columns=['dup_key'])

    # Sort alphabetically by Business Name and Possible Value.
    pvl_df = pvl_df.sort_values(by=['Business Name', 'Possible Value'], key=lambda col: col.str.lower())

    # Remove existing PossibleValuesList tab if exists
    if "PossibleValuesList" in wb.sheetnames:
        wb.remove(wb["PossibleValuesList"])        
        wb.save(xlsx_path)

    # Write new PVL Tab
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pvl_df.to_excel(writer, sheet_name="PossibleValuesList", index=False)
        ws = writer.sheets["PossibleValuesList"]

        max_col = ws.max_column

        # Format the headers and cell widths
        # ------ Header row 1 -------
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col )
        # header1 = ws.cell(row=1, column=1)
        ws['A1'].value = "WSID Possible Values List"
        ws['A1'].font = Font(name="Calibri", size=16,bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="77A7ED", end_color="77A7ED", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")

         # ------ Header row 2 -------
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col )
        # header1 = ws.cell(row=2, column=1)
        ws['A2'].value = "Possible values can be derived by the associated Business Name column in the WSID Details"
        ws['A2'].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        ws['A2'].fill = PatternFill(start_color="77A7ED", end_color="77A7ED", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal="center")

         # ------ Header row 3 -------
        olive_green_fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
        headers = list(pvl_df.columns)
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.font = Font(name="Calibri", size=12,bold=True, color="000000")
            cell.fill = olive_green_fill
            cell.alignment = Alignment(horizontal="center")

        # Format dataa rows (row 4+)
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name="Calibri", size = 8)

        # ------ Auto adjust column widths based on content ------
        for i, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(i)
            for cell in col_cells:
                try:
                   if cell.value:
                       cell_length = len(str(cell.value))
                       if cell_length > max_length:
                           max_length = cell_length
                except:
                    pass       
            ws.column_dimensions[col_letter].width = max_length + 2 # add padding            


        # Access the workbook directly through the writer to set the tab color.
        pvl_ws = writer.book["PossibleValuesList"]
        pvl_ws.sheet_properties.tabColor = "FFFF00"

        # Move the PVL tab to be right after the RD tab.
        sheets = writer.book._sheets
        rd_index = None
        for i, ws  in enumerate(sheets):
            if ws.title == "Resource Details":
                rd_index = i
                break

        if rd_index is not None:
            sheets.remove(pvl_ws)
            sheets.insert(rd_index + 1, pvl_ws)    

        print(f"PossibleValuesList tab generated successfully in {xlsx_path}")    


def convert_api_element_to_name(api_element: str) -> str:
    """
    Convert camelCase API Element to more human-readable business name.
    Example: accountType -> ACOUNT TYPE
    """
    # Split camelCase and capitalized words
    s1 = re.sub('([a-z0-9])([A-Z])', r'\1 \2', api_element)
    return s1.title()


# Load the WSID excel file
xlsx_path = r"C:\Users\A423347\OneDrive - Fidelity Technology Group, LLC\Documents\Brokerage Tax Form Timing\Michael.Montgomery-brokerage-tax_form_timing-1.0.3-resolved.xlsx"

generate_possible_values_list(xlsx_path)

print(f"Processed and saved the updated WSID to {xlsx_path}")
